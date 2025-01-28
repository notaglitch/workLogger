from openpyxl import workbook, load_workbook
from openpyxl.styles import Alignment
import datetime as dt

today_date = dt.datetime.today().strftime('%d/%m/%Y')
wb = load_workbook("Data.xlsx")
ws = wb.active

def find_first_empty_row(ws):
    for row in range(1, ws.max_row + 2):
        if all(cell.value is None for cell in ws[row]):
            return row
    return ws.max_row + 1

first_empty_row = find_first_empty_row(ws)
last_row_value = first_empty_row - 1

last_focus_value = ws[f"B{last_row_value}"].value
last_date_value = ws[f"A{last_row_value}"].value
new_time = int(input("Enter the time in minutes. "))

if last_date_value == today_date:
    last_focus_value += new_time
    ws[f"B{last_row_value}"] = last_focus_value
    print("Updated today's focus")
else:
    ws[f"A{first_empty_row}"] = today_date
    ws[f"B{first_empty_row}"] = new_time
    print("Updated today's date and value")

columns_to_center = ["A", "B"]
for col in columns_to_center:
    for cell in ws[col]:
        cell.alignment = Alignment(horizontal='center')

print(first_empty_row)
wb.save("Data.xlsx")