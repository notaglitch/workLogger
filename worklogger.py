import tkinter as tk
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import datetime as dt
import subprocess
import random
import os

class StopwatchGUI:

    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Worklogger Stopwatch")
        self.root.geometry("400x300")
        self.root.configure(bg='black')
        self.center_window()

        self.is_running = False
        self.elapsed_time = timedelta()
        self.start_time = None

        main_frame = tk.Frame(self.root, bg='black')
        main_frame.pack(expand=True, fill='both', padx=20, pady=20)

        self.time_label = tk.Label(main_frame, text="00:00.00", font=("Courier New", 50), wraplength=400, bg='black', fg='#00ff00')
        self.time_label.pack(expand=True, fill='both', pady=20)

        button_frame = tk.Frame(main_frame, bg='black')
        button_frame.pack(expand=True, fill='x', pady=10)
        button_frame.grid_columnconfigure(0, weight=1)
        button_frame.grid_columnconfigure(1, weight=1)

        self.start_button = tk.Button(button_frame, text="Start", command=self.start_stop, font=("Courier New", 16), width=10, height=2, bg='black', fg='#00ff00', relief='solid', bd=2)
        self.start_button.grid(row=0, column=0, padx=10, pady=10)

        self.finish_button = tk.Button(button_frame, text="Finish", command=self.finish, font=("Courier New", 16), width=10, height=2, bg='black', fg='#00ff00', relief='solid', bd=2)
        self.finish_button.grid(row=0, column=1, padx=10, pady=10)

        self.update_time()

    def center_window(self):
        window_width = 400
        window_height = 300
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        position_top = (screen_height // 2) - (window_height // 2)
        position_right = (screen_width // 2) - (window_width // 2)
        self.root.geometry(f'{window_width}x{window_height}+{position_right}+{position_top}')

    def start_stop(self):
        if self.is_running:
            self.is_running = False
            self.start_button.config(text="Start")
            if self.start_time:
                self.elapsed_time += datetime.now() - self.start_time
        else:
            self.is_running = True
            self.start_button.config(text="Stop")
            self.start_time = datetime.now()

    def finish(self):
        if self.elapsed_time.total_seconds() > 0 or self.is_running:
            current_time = datetime.now()
            time_diff = self.elapsed_time + (current_time - self.start_time) if self.is_running else self.elapsed_time
            minutes = int(time_diff.total_seconds() // 60)
            seconds = int(time_diff.total_seconds() % 60)
            hundredths = int((time_diff.total_seconds() * 100) % 100)

            print(f"Total time: {minutes:02d}:{seconds:02d}.{hundredths:02d}")

            if minutes < 1:
                print("wth, you should work more.")
            else:
                self.update_excel(minutes)

        self.is_running = False
        self.start_button.config(text="Start")
        self.elapsed_time = timedelta()
        self.start_time = None
        self.time_label.config(text="00:00.00")

    def update_excel(self, minutes):
        today_date = dt.datetime.today().strftime('%d/%m/%Y')
                # Use os.path.expanduser to handle iCloud Drive path correctly
        wb_path = os.path.expanduser("~/iCloudDrive/Downloads/Worklogs.xlsx")
        wb = load_workbook(wb_path)
        ws = wb.active

        def find_first_empty_row(ws):
            for row in range(1, ws.max_row + 2):
                if all(cell.value is None for cell in ws[row]):
                    return row
            return ws.max_row + 1

        first_empty_row = find_first_empty_row(ws)
        last_row_value = first_empty_row - 1

        last_focus_value = ws[f"B{last_row_value}"].value
        last_date_value = ws[f"A{last_row_value}"].value
        total_time = minutes

        if last_date_value == today_date:
            last_focus_value += total_time
            ws[f"B{last_row_value}"] = last_focus_value
            print("Updated today's focus")
        else:
            ws[f"A{first_empty_row}"] = today_date
            ws[f"B{first_empty_row}"] = total_time
            print("Updated today's date and value")

        columns_to_center = ["A", "B"]
        for col in columns_to_center:
            for cell in ws[col]:
                cell.alignment = Alignment(horizontal='center')

        wb.save(wb_path)

    def update_time(self):
        if self.is_running:
            current_time = datetime.now()
            time_diff = self.elapsed_time + (current_time - self.start_time)
            minutes = int(time_diff.total_seconds() // 60)
            seconds = int(time_diff.total_seconds() % 60)
            hundredths = int((time_diff.total_seconds() * 100) % 100)

            if random.random() < 0.1:
                seconds = random.randint(0, 59)
                hundredths = random.randint(0, 99)

            time_str = f"{minutes:02d}:{seconds:02d}.{hundredths:02d}"
            self.time_label.config(text=time_str)

        self.root.after(10, self.update_time)

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    stopwatch = StopwatchGUI()
    stopwatch.run()
    subprocess.run(r"excel C:\Users\chotd\iCloudDrive\Downloads\Worklogs.xlsx")
